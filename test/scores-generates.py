import random
import openpyxl
from openpyxl.styles import Font, Alignment

def weighted_random_score():
    """
    Tạo điểm ngẫu nhiên theo tỉ lệ:
    - 0.0 – 5.0:   ~15%
    - 5.1 – 7.0:   ~40%
    - 7.1 – 9.0:   ~30%
    - 9.1 – 10.0:  ~15%
    
    Điểm được làm tròn đến 1 chữ số thập phân.
    """
    r = random.random()
    if r < 0.15:
        score = random.uniform(0, 5)
    elif r < 0.55:   # 0.15 + 0.40
        score = random.uniform(5, 7)
    elif r < 0.85:   # 0.55 + 0.30
        score = random.uniform(7, 9)
    else:
        score = random.uniform(9, 10)
    
    # Làm tròn đến 1 chữ số thập phân
    return round(score, 1)

def main():
    # Nhập số lượng từ người dùng
    try:
        so_luong = int(input("Nhập số lượng hàng dữ liệu (tối đa 1000): "))
        if so_luong > 1000:
            so_luong = 1000
            print("Số lượng vượt quá giới hạn → đặt về 1000.")
        elif so_luong < 1:
            so_luong = 1
            print("Số lượng tối thiểu là 1.")
    except ValueError:
        print("Đầu vào không hợp lệ → sử dụng mặc định 100 hàng.")
        so_luong = 100

    # Tạo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Điểm số"

    # Tiêu đề
    headers = [
        "Điểm quá trình (10%)", "Điểm thành phần (30%)", "Điểm thi (60%)"
    ]
    ws['A1'] = headers[0]
    ws['B1'] = headers[1]
    ws['C1'] = headers[2]

    # Định dạng header
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}1']
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Tạo dữ liệu
    for row in range(2, so_luong + 2):
        diem_qua_trinh = weighted_random_score()
        diem_thanh_phan = weighted_random_score()
        diem_thi       = weighted_random_score()

        ws[f'A{row}'] = diem_qua_trinh
        ws[f'B{row}'] = diem_thanh_phan
        ws[f'C{row}'] = diem_thi

        # Căn giữa và định dạng số (hiển thị 1 chữ số thập phân)
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.alignment = Alignment(horizontal='center')
            # Buộc hiển thị 1 chữ số thập phân (ví dụ: 7.0 thay vì 7)
            cell.number_format = '0.0'

    # Tự động điều chỉnh độ rộng cột
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 5
        ws.column_dimensions[column_letter].width = adjusted_width

    # Lưu file
    file_name = "diem_so_1_chu_so_thap_phan.xlsx"
    wb.save(file_name)
    print(f"\nĐã tạo file '{file_name}' thành công!")
    print(f"Tổng số hàng dữ liệu: {so_luong}")
    print("Tất cả điểm số đã được làm tròn đến 1 chữ số thập phân.")
    print("Phân bố điểm: ~15% (0.0-5.0) - 40% (5.1-7.0) - 30% (7.1-9.0) - 15% (9.1-10.0)")

if __name__ == "__main__":
    main()