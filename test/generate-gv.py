import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import random
import re
import unidecode  # <-- thêm thư viện này để bỏ dấu dễ dàng

# Tạo workbook mới
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Giảng viên"

# Định nghĩa màu sắc và border
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Danh sách môn học (giữ nguyên)
mon_hoc = [
    ("ML1011", "Triết học Mác - Lênin"),
    ("TH1001", "Giáo dục thể chất 1"),
    ("TN1001", "Tiếng Anh 1"),
    ("TH1101", "Toán cao cấp A1"),
    ("TH1102", "Đại số tuyến tính"),
    ("IT1001", "Nhập môn lập trình"),
    ("IT1002", "Tin học cơ sở"),
    ("ML1021", "Kinh tế chính trị Mác - Lênin"),
    ("TH1002", "Giáo dục thể chất 2"),
    ("TN1002", "Tiếng Anh 2"),
    ("TH1103", "Toán cao cấp A2"),
    ("TH1104", "Toán rời rạc"),
    ("IT1003", "Lập trình hướng đối tượng"),
    ("IT1004", "Cấu trúc dữ liệu và giải thuật"),
    ("ML1031", "Chủ nghĩa xã hội khoa học"),
    ("TH1003", "Giáo dục thể chất 3"),
    ("TN1003", "Tiếng Anh 3"),
    ("TH1105", "Xác suất thống kê"),
    ("IT2001", "Cơ sở dữ liệu"),
    ("IT2002", "Hệ điều hành"),
    ("IT2003", "Kiến trúc máy tính"),
    ("ML1041", "Tư tưởng Hồ Chí Minh"),
    ("PL1001", "Pháp luật đại cương"),
    ("IT2004", "Mạng máy tính"),
    ("IT2005", "Lập trình Web"),
    ("IT2006", "Phân tích thiết kế hệ thống"),
    ("IT2007", "Công nghệ phần mềm"),
    ("IT2008", "Trí tuệ nhân tạo"),
    ("ML1051", "Lịch sử Đảng Cộng sản Việt Nam"),
    ("IT3001", "Lập trình di động"),
    ("IT3002", "An toàn và bảo mật thông tin"),
    ("IT3003", "Học máy (Machine Learning)"),
    ("IT3004", "Lập trình Java nâng cao"),
    ("IT3005", "Đồ họa máy tính"),
    ("IT3006", "Phát triển ứng dụng thương mại điện tử"),
    ("IT3007", "Quản trị mạng"),
    ("IT3008", "Hệ quản trị cơ sở dữ liệu"),
    ("IT3009", "Điện toán đám mây (Cloud Computing)"),
    ("IT3010", "Lập trình Python"),
    ("IT3011", "Phân tích dữ liệu lớn (Big Data)"),
    ("IT3012", "Internet of Things (IoT)"),
    ("IT3013", "Blockchain và ứng dụng"),
    ("IT4001", "Phát triển game"),
    ("IT4002", "Xử lý ngôn ngữ tự nhiên (NLP)"),
    ("IT4003", "Computer Vision"),
    ("IT4004", "DevOps"),
    ("IT4005", "Quản lý dự án CNTT"),
    ("IT4006", "Kiểm thử phần mềm"),
    ("IT4007", "Thực tập doanh nghiệp"),
    ("IT4008", "Đồ án chuyên ngành"),
    ("IT4009", "Khóa luận tốt nghiệp"),
    ("IT4010", "Seminar chuyên ngành"),
]

# Dữ liệu mẫu họ tên
ho_dem = ["Nguyễn", "Trần", "Lê", "Phạm", "Hoàng", "Phan", "Vũ", "Đặng", "Bùi", "Đỗ", "Hồ", "Ngô", "Dương", "Lý"]
ten_dem = ["Văn", "Thị", "Hữu", "Đức", "Minh", "Anh", "Tuấn", "Quốc", "Thanh", "Ngọc", "Hồng", "Phương"]
ten = [
    "An", "Bình", "Cường", "Dũng", "Hải", "Hòa", "Hùng", "Khanh",
    "Linh", "Long", "Minh", "Nam", "Phong", "Phúc", "Quân", "Sơn",
    "Thắng", "Thành", "Thiện", "Tiến", "Toàn", "Trang", "Trung",
    "Tuấn", "Việt", "Vinh", "Yến", "Ngọc", "Hương", "Mai", "Lan"
]
dia_chi = ["Hà Nội", "TP. Hồ Chí Minh", "Đà Nẵng", "Hải Phòng", "Cần Thơ", "Huế", "Nha Trang", "Vũng Tàu", "Biên Hòa", "Thái Nguyên"]

# Tạo header
headers = ['STT', 'Mã giảng viên', 'Họ tên', 'Ngày sinh', 'Email', 'Số điện thoại', 'Giới tính', 'Địa chỉ']
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = header_fill
    cell.border = thin_border

ws.row_dimensions[1].height = 25

# Hàm tạo mã giảng viên dạng TRAN_TRUNG_CUONG
def tao_ma_giang_vien(ho_ten):
    # Bỏ dấu → viết HOA → thay khoảng trắng bằng _
    khong_dau = unidecode.unidecode(ho_ten).upper()
    ma = khong_dau.replace(" ", "_")
    return ma

# Hàm tạo email không dấu, chữ thường
def tao_email_khong_dau(ho, td, t, index):
    # Nối họ + tên đệm + tên → bỏ dấu → chữ thường
    full = f"{ho} {td} {t}"
    khong_dau = unidecode.unidecode(full).lower().replace(" ", "")
    email = f"{khong_dau}{index}@university.edu.vn"
    return email

# Tạo dữ liệu 30 giảng viên
giang_vien_data = []
used_emails = set()
used_phones = set()

for i in range(1, 31):
    # Tạo tên ngẫu nhiên
    ho = random.choice(ho_dem)
    td = random.choice(ten_dem)
    t = random.choice(ten)
    ho_ten = f"{ho} {td} {t}"
    
    # Tạo mã giảng viên mới theo dạng yêu cầu
    ma_gv = tao_ma_giang_vien(ho_ten)
    
    # Ngày sinh (1970–1990)
    start_date = datetime(1970, 1, 1)
    end_date = datetime(1990, 12, 31)
    days_between = (end_date - start_date).days
    random_days = random.randint(0, days_between)
    ngay_sinh = start_date + timedelta(days=random_days)
    
    # Email không dấu, unique
    email = tao_email_khong_dau(ho, td, t, i)
    base = email.split('@')[0]
    counter = 0
    while email in used_emails:
        counter += 1
        email = f"{base}{counter}@university.edu.vn"
    used_emails.add(email)
    
    # Số điện thoại unique
    phone = f"09{random.randint(10000000, 99999999)}"
    while phone in used_phones:
        phone = f"09{random.randint(10000000, 99999999)}"
    used_phones.add(phone)
    
    # Giới tính
    gioi_tinh = random.choice(["NAM", "NU"])
    
    # Địa chỉ
    dia_chi_gv = random.choice(dia_chi)
    
    giang_vien_data.append({
        'stt': i,
        'ma_gv': ma_gv,
        'ho_ten': ho_ten,
        'ngay_sinh': ngay_sinh.strftime('%Y-%m-%d'),
        'email': email,
        'sdt': phone,
        'gioi_tinh': gioi_tinh,
        'dia_chi': dia_chi_gv,
        'mon_day': []
    })

# Phân công môn học (giữ nguyên logic)
mon_hoc_shuffled = mon_hoc.copy()
random.shuffle(mon_hoc_shuffled)

idx = 0
for gv in giang_vien_data:
    num_subjects = random.randint(2, 5)
    for _ in range(num_subjects):
        if idx < len(mon_hoc_shuffled):
            gv['mon_day'].append(mon_hoc_shuffled[idx][0])
            idx += 1
        else:
            break

# Ghi dữ liệu vào sheet "Giảng viên" (giữ nguyên phần này)
current_row = 2
for gv in giang_vien_data:
    ws.cell(row=current_row, column=1).value = gv['stt']
    ws.cell(row=current_row, column=2).value = gv['ma_gv']
    ws.cell(row=current_row, column=3).value = gv['ho_ten']
    ws.cell(row=current_row, column=4).value = gv['ngay_sinh']
    ws.cell(row=current_row, column=5).value = gv['email']
    ws.cell(row=current_row, column=6).value = gv['sdt']
    ws.cell(row=current_row, column=7).value = gv['gioi_tinh']
    ws.cell(row=current_row, column=8).value = gv['dia_chi']

    for col in range(1, 9):
        cell = ws.cell(row=current_row, column=col)
        cell.alignment = Alignment(horizontal='center' if col in [1,2,4,6,7] else 'left', vertical='center')
        cell.border = thin_border

    ws.row_dimensions[current_row].height = 20
    current_row += 1

# Định dạng cột (giữ nguyên)
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 22   # rộng hơn tí vì mã dài
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 38
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 25

# Sheet phân công môn học (giữ nguyên)
ws2 = wb.create_sheet(title="Phân công môn học")

headers2 = ['Mã giảng viên', 'Mã môn học']
for col, header in enumerate(headers2, start=1):
    cell = ws2.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = header_fill
    cell.border = thin_border

current_row = 2
for gv in giang_vien_data:
    for ma_mon in gv['mon_day']:
        ws2.cell(row=current_row, column=1).value = gv['ma_gv']
        ws2.cell(row=current_row, column=2).value = ma_mon
        for col in [1,2]:
            cell = ws2.cell(row=current_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws2.row_dimensions[current_row].height = 20
        current_row += 1

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 15

# Lưu file
wb.save('giang_vien_va_phan_cong_mon_hoc.xlsx')
print("Đã tạo file thành công!")